import openpyxl
workbook = openpyxl.Workbook()

# Создаем новый лист
worksheet = workbook.active

# Создаем массив значениями
my_data = [1,2,3,4,5]
worksheet.cell(row=1,column=1,value='Саша')
worksheet.cell(row=1,column=4,value='Рома')
# Заполняем таблицу данными из массива
for i in range(len(my_data)):
    # Выставляем значения на пересечении строки и столбца
    worksheet.cell(row=i+2,column=1,value=my_data[i])

# Сохраняем файл
workbook.save(filename = 'my_file.xlsx')